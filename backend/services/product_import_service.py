"""ProductImportService — human-reviewed bulk product import from .xlsx.

Two phases over the SAME file (stateless):
  * validate (commit=False): parse + validate every row, return a per-row preview.
  * import   (commit=True):  create only the VALID rows via the existing
    ProductService (no business logic bypassed; no inventory writes; no overwrite
    of existing products).

Existing products with a matching name are flagged as duplicates and skipped.
"""
from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from models.product import Product
from models.user import User
from schemas.product import ProductCreate
from schemas.product_import import (
    ImportDuplicate, ImportError, ImportResult, ImportRow,
)
from utils.enums import ProcurementMethod
from utils.exceptions import ValidationError

# Canonical headers (matched case-insensitively, order-independent).
COLUMNS = ["Product Name", "Sales Price", "Cost Price",
           "Procurement Route", "Procure On Demand", "Default Vendor ID"]
REQUIRED_HEADERS = COLUMNS[:-1]  # vendor id is optional

_ROUTES = {"purchase": ProcurementMethod.PURCHASE,
           "manufacture": ProcurementMethod.MANUFACTURE}
_BOOLS = {"yes": True, "true": True, "no": False, "false": False}

SAMPLE_ROWS = [
    ("Wooden Table", 120, 40, "Manufacture", "Yes", None),
    ("Office Chair", 80, 30, "Purchase", "Yes", 3),
    ("Phone", 2000, 1000, "Purchase", "Yes", 2),
]


class _RowError(Exception):
    def __init__(self, field: str, message: str):
        self.field = field
        self.message = message


class ProductImportService:
    def __init__(self, db: Session):
        self.db = db

    # ---- downloadable template ----
    def build_template(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(COLUMNS)
        for r in SAMPLE_ROWS:
            ws.append(list(r))
        for i, _ in enumerate(COLUMNS, start=1):
            ws.column_dimensions[chr(64 + i)].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ---- export current catalogue ----
    def build_export(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(COLUMNS)
        for p in self.db.query(Product).order_by(Product.id).all():
            ws.append([
                p.name, float(p.sales_price), float(p.cost_price),
                p.procurement_method.value.title(),
                "Yes" if p.procure_on_demand else "No",
                p.vendor_id,
            ])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ---- validate + (optionally) import ----
    def process(self, content: bytes, *, commit: bool, user_id: int) -> ImportResult:
        ws = self._open(content)
        col = self._header_map(ws)

        existing = {n.lower() for (n,) in self.db.query(Product.name).all()}
        seen: set[str] = set()  # names accepted earlier in THIS file

        rows: list[ImportRow] = []
        duplicates: list[ImportDuplicate] = []
        errors: list[ImportError] = []
        success = 0

        for r in range(2, ws.max_row + 1):
            values = {key: ws.cell(row=r, column=idx).value for key, idx in col.items()}
            if all(v is None or str(v).strip() == "" for v in values.values()):
                continue  # skip fully blank rows

            name = str(values.get("Product Name") or "").strip()
            try:
                payload = self._parse_row(values)
            except _RowError as e:
                errors.append(ImportError(row=r, field=e.field, message=e.message))
                rows.append(ImportRow(row=r, product_name=name, status="error", reason=e.message))
                continue

            key = payload.name.lower()
            if key in existing or key in seen:
                duplicates.append(ImportDuplicate(row=r, product_name=payload.name))
                rows.append(ImportRow(row=r, product_name=payload.name, status="duplicate",
                                      reason="Product name already exists — skipped."))
                continue

            seen.add(key)
            if commit:
                from services.product_service import ProductService
                try:
                    ProductService(self.db).create(payload, user_id=user_id)
                except Exception as exc:  # surface unexpected create failure as a row error
                    errors.append(ImportError(row=r, field="product_name", message=str(exc)))
                    rows.append(ImportRow(row=r, product_name=payload.name, status="error", reason=str(exc)))
                    continue
                rows.append(ImportRow(row=r, product_name=payload.name, status="created"))
            else:
                rows.append(ImportRow(row=r, product_name=payload.name, status="valid"))
            success += 1

        return ImportResult(
            success_count=success,
            failure_count=len(errors),
            duplicate_count=len(duplicates),
            duplicates=duplicates,
            errors=errors,
            rows=rows,
            committed=commit,
        )

    # ---- helpers ----
    def _open(self, content: bytes):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            return wb.active
        except Exception as exc:
            raise ValidationError(f"Could not read the Excel file: {exc}") from exc

    def _header_map(self, ws) -> dict[str, int]:
        header = {}
        for idx in range(1, (ws.max_column or 0) + 1):
            raw = ws.cell(row=1, column=idx).value
            if raw is None:
                continue
            header[str(raw).strip().lower()] = idx
        col = {}
        for c in COLUMNS:
            if c.lower() in header:
                col[c] = header[c.lower()]
        missing = [c for c in REQUIRED_HEADERS if c not in col]
        if missing:
            raise ValidationError(
                "Missing required column(s): " + ", ".join(missing)
                + ". Download the template for the correct format."
            )
        return col

    def _parse_row(self, values: dict) -> ProductCreate:
        name = str(values.get("Product Name") or "").strip()
        if not name:
            raise _RowError("product_name", "Product name is required.")

        sales_price = self._num(values.get("Sales Price"), "sales_price", "Sales Price")
        cost_price = self._num(values.get("Cost Price"), "cost_price", "Cost Price")

        route_raw = str(values.get("Procurement Route") or "").strip().lower()
        route = _ROUTES.get(route_raw)
        if route is None:
            raise _RowError("procurement_route", "Must be 'Purchase' or 'Manufacture'.")

        pod_raw = str(values.get("Procure On Demand") or "").strip().lower()
        if pod_raw not in _BOOLS:
            raise _RowError("procure_on_demand", "Must be Yes/No (or True/False).")
        procure_on_demand = _BOOLS[pod_raw]

        vendor_id = self._vendor(values.get("Default Vendor ID"))

        return ProductCreate(
            name=name, sales_price=sales_price, cost_price=cost_price,
            on_hand_qty=0,  # import never sets inventory
            procurement_method=route, procure_on_demand=procure_on_demand,
            vendor_id=vendor_id,
        )

    @staticmethod
    def _num(value, field: str, label: str) -> float:
        if value is None or str(value).strip() == "":
            raise _RowError(field, f"{label} is required.")
        try:
            n = float(value)
        except (TypeError, ValueError):
            raise _RowError(field, "Invalid numeric value.")
        if n < 0:
            raise _RowError(field, "Must be greater than or equal to 0.")
        return n

    def _vendor(self, value) -> int | None:
        if value is None or str(value).strip() == "":
            return None
        try:
            vid = int(float(value))
        except (TypeError, ValueError):
            raise _RowError("default_vendor_id", "Vendor ID must be a number.")
        if self.db.get(User, vid) is None:
            raise _RowError("default_vendor_id", f"Vendor {vid} does not exist.")
        return vid
